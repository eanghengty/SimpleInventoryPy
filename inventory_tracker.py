import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import Table, TableStyle, SimpleDocTemplate
from reportlab.lib.units import inch

file_name = "inventory.xlsx"

# Create file if it doesn't exist
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product ID", "Product Name", "Stock","Unit", "Last Updated", "Status"])
    wb.save(file_name)
#just add on 19-Apr-2025
def export_inventory_report_pdf():
    wb,ws=load_inventory()
    today= datetime.today().strftime("%Y-%m-%d")
    filename=f"inventory_report_{today.replace("-","")}.pdf"

    data = [["Product ID","Product Name","Stock","Unit","Last Updated","Status"]]
    for row in ws.iter_rows(min_row=2,values_only=True):
        pid,name,stock,unit,date,status=row
        status="LOW" if stock < 5 else "OK"

        data.append([pid,name,stock,unit,date,status])

    pdf=SimpleDocTemplate(filename,pagesize=A4)
    table=Table(data,repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.lightblue),
        ('GRID',(0,0),(-1,-1),0.5,colors.grey),
        ('TEXTCOLOR',(0,0),(-1,0),colors.black),
        ('ALIGN',(2,1),(2,-1),'RIGHT'),
        ('BACKGROUND',(0,1),(-1,-1), colors.whitesmoke),
    ]))
    elements=[table]
    pdf.build(elements)

    print(f"\n✅ PDF report exported as: {filename}")

def load_inventory():
    wb = load_workbook(file_name)
    ws = wb.active
    return wb, ws

def edit_product_info():
    pid=input("Enter product ID to edit: ").strip()
    wb,ws=load_inventory()
    found=False

    for row in ws.iter_rows(min_row=2):
        if row[0].value==pid:
            print(f"\nEditing product: {row[1].value}")
            new_name=input(f"Enter new name (or press enter to keep '{row[1].value}'): ").strip()
            new_unit=input(f"\nEnter new unit (or press enter to keep '{row[3].value}'): ").strip()
            new_stock_input=input(f"Enter new stock (or press enter to keep '{row[2].value}'): ").strip()

            if new_name:
                row[1].value= new_name
            if new_unit:
                row[3].value= new_unit
            if new_stock_input:
                try:
                    new_stock = int(new_stock_input)
                    row[2].value = new_stock
                    row[5].value = "LOW" if new_stock < 5 else "OK"
                except ValueError:
                    print("Invalid stock input. Keeping previous value.")

            row[4].value = datetime.today().strftime("%Y-%m-%d")
            found = True
            break
    
    if found:
        wb.save(file_name)
        print("✅ Product updated.")

    else:
        print("❌ Product ID not found.")

def add_product():
    pid = input("Enter Product ID: ").strip()
    name = input("Enter Product Name: ").strip()
    stock = int(input("Enter initial stock: "))
    unit = input("Enter unit (pcs, box, kg, etc.): ").strip()
    date = datetime.today().strftime("%Y-%m-%d")
    status = "LOW" if stock < 5 else "OK"
    wb, ws = load_inventory()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == pid:
            print("Product ID already exists.")
            return

    ws.append([pid, name, stock,unit, date, status])
    wb.save(file_name)
    print("Product added.")

def update_stock():
    pid = input("Enter Product ID to update: ").strip()
    change = int(input("Enter stock change (positive for add, negative for remove): "))
    date = datetime.today().strftime("%Y-%m-%d")

    wb, ws = load_inventory()

    log_ws = wb["Logs"] if "Logs" in wb.sheetnames else wb.create_sheet("Logs")
    if log_ws.max_row==1:
        log_ws.append(["Date","Product ID","Change","New Stock","Action"])

    found = False

    for row in ws.iter_rows(min_row=2):
        if row[0].value == pid:
            row[2].value += change
            row[4].value=date
            row[5].value = "LOW" if row[2].value<5 else "OK"
            new_stock=row[2].value
            action="Stock In" if change > 0 else "Stock Out"
            log_ws.append([date,pid,change,new_stock,action])
            found = True
            break

    if found:
        wb.save(file_name)
        print("Stock updated.")
    else:
        print("Product not found.")

def delete_product():
    pid=input("Enter Product ID to delete: ").strip()
    wb,ws=load_inventory()
    found=False

    for row in ws.iter_rows(min_row=2):
        if row[0].value == pid:
            current_stock=row[2].value
            if current_stock>0:
                print(f"❌ Cannot delete product '{pid}' — stock remaining: {current_stock}")
                return #exit without deleting
            
            confirm=input(f"Are you sure you want to delete '{row[1].value} - {row[2].value}'? (yes/no): ").strip()
            if confirm == "yes":
                ws.delete_rows(row[0].row,1)
                print(f"✅ Product '{pid}' deleted.")
                found=True

                #log deletion
                log_ws=wb['Logs'] if "Logs" in wb.sheetnames else wb.create_sheet("Logs")
                if log_ws.max_row==1:
                    log_ws.append(["Date","Product ID","Changes","New Stock", "Action"])
                log_ws.append([datetime.today().strftime("%Y-%m-%d"),pid,"-","-","Deleted"])
                break
            else:
                print("❌ Deletion cancelled.")
                return
    if found:
        wb.save(file_name)
    else:
        print("❌ Product not found.")

def show_inventory():
    wb, ws = load_inventory()
    print("\nCurrent Inventory:")
    print("{:<10} {:<20} {:<10} {:<6} {:<15} {:<10}".format("ID", "Name", "Stock","Unit", "Last Updated", "Status"))
    for row in ws.iter_rows(min_row=2, values_only=True):
        pid,name,stock,unit,date,status=row
        status = "LOW" if stock < 5 else "OK"
        print("{:<10} {:<20} {:<10} {:<6} {:<15} {:<10}".format(pid,name,stock,unit,date,status))

    print("\n* Products marked 'LOW' need restocking.")

def menu():
    while True:
        print("\nInventory Tracker")
        print("1. Add Product")
        print("2. Update Stock")
        print("3. View Inventory")
        print("4. Export Inventory Report (PDF)")
        print("5. Edit Product Info")
        print("6. Delete Product")
        print("7. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            add_product()
        elif choice == "2":
            update_stock()
        elif choice == "3":
            show_inventory()
        elif choice == "4":
            export_inventory_report_pdf()
        elif choice == "5":
            edit_product_info()
        elif choice == "6":
            delete_product()
        elif choice == "7":
            break
        else:
            print("Invalid option.")

if __name__ == "__main__":
    menu()
