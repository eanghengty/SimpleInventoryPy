import openpyxl
from openpyxl import load_workbook
from datetime import datetime
import os

file_name = "inventory.xlsx"

# Create file if it doesn't exist
if not os.path.exists(file_name):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Product ID", "Product Name", "Stock", "Last Updated"])
    wb.save(file_name)

def load_inventory():
    wb = load_workbook(file_name)
    ws = wb.active
    return wb, ws

def add_product():
    pid = input("Enter Product ID: ").strip()
    name = input("Enter Product Name: ").strip()
    stock = int(input("Enter initial stock: "))
    date = datetime.today().strftime("%Y-%m-%d")

    wb, ws = load_inventory()

    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] == pid:
            print("Product ID already exists.")
            return

    ws.append([pid, name, stock, date])
    wb.save(file_name)
    print("Product added.")

def update_stock():
    pid = input("Enter Product ID to update: ").strip()
    change = int(input("Enter stock change (positive for add, negative for remove): "))

    wb, ws = load_inventory()
    found = False

    for row in ws.iter_rows(min_row=2):
        if row[0].value == pid:
            row[2].value += change
            row[3].value = datetime.today().strftime("%Y-%m-%d")
            found = True
            break

    if found:
        wb.save(file_name)
        print("Stock updated.")
    else:
        print("Product not found.")

def show_inventory():
    wb, ws = load_inventory()
    print("\nCurrent Inventory:")
    print("{:<10} {:<20} {:<10} {:<15}".format("ID", "Name", "Stock", "Last Updated"))
    for row in ws.iter_rows(min_row=2, values_only=True):
        print("{:<10} {:<20} {:<10} {:<15}".format(*row))

def menu():
    while True:
        print("\nInventory Tracker")
        print("1. Add Product")
        print("2. Update Stock")
        print("3. View Inventory")
        print("4. Exit")
        choice = input("Choose an option: ")

        if choice == "1":
            add_product()
        elif choice == "2":
            update_stock()
        elif choice == "3":
            show_inventory()
        elif choice == "4":
            break
        else:
            print("Invalid option.")

if __name__ == "__main__":
    menu()
